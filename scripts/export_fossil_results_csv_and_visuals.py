#!/usr/bin/env python3
from __future__ import annotations

import csv
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "outputs" / "fossil_experiment_results.xlsx"
EXPORT_ROOT = ROOT / "outputs" / "fossil_results_exports"
CSV_DIR = EXPORT_ROOT / "csv"
VISUALS_DIR = EXPORT_ROOT / "visuals"
CHART_SOURCE_DIR = ROOT / "outputs" / "fossil_results_charts"

SHEET_EXPORTS = {
    "All Tests": "all_tests.csv",
    "Model Comparison": "model_comparison_initial_vs_updated.csv",
    "Per Class Updated": "per_class_updated_38_image_test.csv",
    "Incorrect Updated": "incorrect_predictions_updated_38_image_test.csv",
    "Training History": "training_history_all_models.csv",
    "Notes": "notes.csv",
}

VISUAL_EXPORTS = {
    "accuracy_comparison.png": "01_accuracy_initial_vs_updated.png",
    "correct_vs_incorrect.png": "02_correct_vs_incorrect_updated_38.png",
    "per_class_accuracy.png": "03_per_class_accuracy_updated_38.png",
}


def normalize(value):
    if value is None:
        return ""
    return value


def export_sheet(ws, output_path: Path) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([normalize(value) for value in row])


def main() -> None:
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    VISUALS_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(WORKBOOK, data_only=True)
    for sheet_name, filename in SHEET_EXPORTS.items():
        export_sheet(wb[sheet_name], CSV_DIR / filename)

    for source_name, export_name in VISUAL_EXPORTS.items():
        shutil.copy2(CHART_SOURCE_DIR / source_name, VISUALS_DIR / export_name)

    readme = EXPORT_ROOT / "README_RESULTS_INDEX.md"
    readme.write_text(
        """# Fossil Results Export Index

## CSV Files

- `csv/all_tests.csv`: every recorded test run, including exploratory tests and preprocessing experiment tests.
- `csv/model_comparison_initial_vs_updated.csv`: main comparison table for the five preprocessing models, showing initial 10-image test versus updated 38-image test.
- `csv/per_class_updated_38_image_test.csv`: coral-vs-shell performance for each model on the updated 38-image wild test.
- `csv/incorrect_predictions_updated_38_image_test.csv`: every wrong prediction on the updated 38-image wild test, with true class, predicted class, and confidence.
- `csv/training_history_all_models.csv`: epoch-by-epoch training and validation metrics for all saved models.
- `csv/notes.csv`: definitions and interpretation notes.

## Visuals

- `visuals/01_accuracy_initial_vs_updated.png`: bar chart comparing initial wild-test accuracy against updated 38-image wild-test accuracy.
- `visuals/02_correct_vs_incorrect_updated_38.png`: stacked bar chart showing correct versus incorrect predictions on the updated test set.
- `visuals/03_per_class_accuracy_updated_38.png`: line chart comparing coral and shell accuracy per model on the updated test set.

## Key Interpretation

- The updated 38-image wild test is more reliable than the initial 10-image wild test.
- Current best model on the updated test is `original_enhanced`: 34/38 correct, 89.5% accuracy.
- The `original_enhanced` model was trained on original plus enhanced JPEG versions, not enhanced-only images.
""",
        encoding="utf-8",
    )

    print(f"Exported CSVs to {CSV_DIR}")
    print(f"Exported visuals to {VISUALS_DIR}")
    print(f"Wrote index to {readme}")


if __name__ == "__main__":
    main()
