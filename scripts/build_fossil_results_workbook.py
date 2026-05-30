#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
EXPLORATORY = DATA / "exploratory"
EXPERIMENTS = DATA / "preprocessing_experiments"
OUTPUT_DIR = ROOT / "outputs"
ASSET_DIR = OUTPUT_DIR / "fossil_results_charts"
OUTPUT_XLSX = OUTPUT_DIR / "fossil_experiment_results.xlsx"

EXPERIMENT_ORDER = [
    "original_only",
    "original_enhanced",
    "original_grayscale",
    "original_clahe",
    "all_variants",
]

DISPLAY_NAMES = {
    "original_only": "Original only",
    "original_enhanced": "Original + enhanced",
    "original_grayscale": "Original + grayscale",
    "original_clahe": "Original + CLAHE",
    "all_variants": "All variants",
    "exploratory_original_enhanced": "Exploratory: original + enhanced test",
    "exploratory_original_only": "Exploratory: original-only test",
}


def read_json(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def pct(value: float) -> float:
    return round(float(value), 4)


def confusion_metrics(summary: dict) -> dict:
    matrix = summary.get("confusion_matrix", [[0, 0], [0, 0]])
    coral_total = sum(matrix[0])
    shell_total = sum(matrix[1])
    coral_correct = matrix[0][0] if coral_total else 0
    shell_correct = matrix[1][1] if shell_total else 0
    return {
        "coral_correct": coral_correct,
        "coral_total": coral_total,
        "coral_accuracy": coral_correct / coral_total if coral_total else None,
        "shell_correct": shell_correct,
        "shell_total": shell_total,
        "shell_accuracy": shell_correct / shell_total if shell_total else None,
        "coral_as_shell": matrix[0][1] if len(matrix[0]) > 1 else 0,
        "shell_as_coral": matrix[1][0] if len(matrix) > 1 else 0,
    }


def collect_results() -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    all_tests: list[dict] = []
    preprocessing_comparison: list[dict] = []
    per_class: list[dict] = []
    incorrect_predictions: list[dict] = []

    exploratory_tests = [
        (
            "exploratory_original_enhanced",
            "Exploratory baseline",
            "Initial 10 wild images, tested as original + enhanced variants",
            EXPLORATORY / "wild_test" / "wild_summary.json",
        ),
        (
            "exploratory_original_only",
            "Exploratory baseline",
            "Initial 10 wild images, original images only",
            EXPLORATORY / "wild_test_original_only" / "wild_summary.json",
        ),
    ]
    for test_id, family, test_set, path in exploratory_tests:
        if not path.exists():
            continue
        summary = read_json(path)
        all_tests.append(
            {
                "test_id": test_id,
                "display_name": DISPLAY_NAMES[test_id],
                "family": family,
                "train_variants": "original+enhanced",
                "test_set": test_set,
                "total": summary["total"],
                "correct": summary["correct"],
                "incorrect": summary["total"] - summary["correct"],
                "accuracy": summary["accuracy"],
                "source_file": str(path.relative_to(ROOT)),
            }
        )

    old_rows = {
        row["experiment"]: row
        for row in read_csv(EXPERIMENTS / "preprocessing_experiment_summary.csv")
    }
    for exp in EXPERIMENT_ORDER:
        old = old_rows[exp]
        old_total = int(old["wild_total"])
        old_correct = int(old["wild_correct"])
        old_acc = float(old["wild_accuracy"])
        new_path = EXPERIMENTS / exp / "wild_test_updated_38_no_torchvision" / "wild_summary.json"
        new = read_json(new_path)

        preprocessing_comparison.append(
            {
                "experiment": exp,
                "display_name": DISPLAY_NAMES[exp],
                "train_variants": old["train_variants"],
                "initial_correct": old_correct,
                "initial_total": old_total,
                "initial_incorrect": old_total - old_correct,
                "initial_accuracy": old_acc,
                "updated_correct": new["correct"],
                "updated_total": new["total"],
                "updated_incorrect": new["total"] - new["correct"],
                "updated_accuracy": new["accuracy"],
                "accuracy_change": new["accuracy"] - old_acc,
            }
        )

        all_tests.append(
            {
                "test_id": f"{exp}_initial_10",
                "display_name": DISPLAY_NAMES[exp],
                "family": "Preprocessing shootout",
                "train_variants": old["train_variants"],
                "test_set": "Initial 10 wild images, original images only",
                "total": old_total,
                "correct": old_correct,
                "incorrect": old_total - old_correct,
                "accuracy": old_acc,
                "source_file": "data/preprocessing_experiments/preprocessing_experiment_summary.csv",
            }
        )
        all_tests.append(
            {
                "test_id": f"{exp}_updated_38",
                "display_name": DISPLAY_NAMES[exp],
                "family": "Preprocessing shootout",
                "train_variants": old["train_variants"],
                "test_set": "Updated 38-image wild test",
                "total": new["total"],
                "correct": new["correct"],
                "incorrect": new["total"] - new["correct"],
                "accuracy": new["accuracy"],
                "source_file": str(new_path.relative_to(ROOT)),
            }
        )

        metrics = confusion_metrics(new)
        per_class.append(
            {
                "experiment": DISPLAY_NAMES[exp],
                "coral_correct": metrics["coral_correct"],
                "coral_total": metrics["coral_total"],
                "coral_accuracy": metrics["coral_accuracy"],
                "shell_correct": metrics["shell_correct"],
                "shell_total": metrics["shell_total"],
                "shell_accuracy": metrics["shell_accuracy"],
                "coral_as_shell": metrics["coral_as_shell"],
                "shell_as_coral": metrics["shell_as_coral"],
            }
        )

        pred_path = EXPERIMENTS / exp / "wild_test_updated_38_no_torchvision" / "wild_predictions.csv"
        for row in read_csv(pred_path):
            if row["correct"] == "False":
                incorrect_predictions.append(
                    {
                        "experiment": DISPLAY_NAMES[exp],
                        "image": Path(row["image"]).name,
                        "folder": Path(row["image"]).parent.name,
                        "true_class": row["true_class"],
                        "predicted_class": row["predicted_class"],
                        "confidence": float(row["confidence"]),
                    }
                )

    return all_tests, preprocessing_comparison, per_class, incorrect_predictions


def collect_training_history() -> list[dict]:
    rows: list[dict] = []
    paths = [(EXPLORATORY / "model" / "training_history.json", "Exploratory baseline")]
    for exp in EXPERIMENT_ORDER:
        paths.append((EXPERIMENTS / exp / "model" / "training_history.json", DISPLAY_NAMES[exp]))

    for path, name in paths:
        if not path.exists():
            continue
        for row in read_json(path):
            rows.append(
                {
                    "model": name,
                    "epoch": row["epoch"],
                    "train_accuracy": row["train_accuracy"],
                    "val_accuracy": row["val_accuracy"],
                    "train_loss": row["train_loss"],
                    "val_loss": row["val_loss"],
                }
            )
    return rows


def make_charts(preprocessing: list[dict], per_class: list[dict]) -> dict[str, Path]:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    labels = [r["display_name"] for r in preprocessing]
    initial = [r["initial_accuracy"] * 100 for r in preprocessing]
    updated = [r["updated_accuracy"] * 100 for r in preprocessing]
    correct = [r["updated_correct"] for r in preprocessing]
    incorrect = [r["updated_incorrect"] for r in preprocessing]

    plt.style.use("seaborn-v0_8-whitegrid")

    fig, ax = plt.subplots(figsize=(10.5, 5.5))
    x = range(len(labels))
    width = 0.36
    ax.bar([i - width / 2 for i in x], initial, width, label="Initial 10-image test", color="#7aa6c2")
    ax.bar([i + width / 2 for i in x], updated, width, label="Updated 38-image test", color="#2f6f73")
    ax.set_ylabel("Accuracy (%)")
    ax.set_title("Wild Test Accuracy: Initial vs Updated Test Set")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=25, ha="right")
    ax.set_ylim(0, 105)
    ax.legend()
    fig.tight_layout()
    accuracy_path = ASSET_DIR / "accuracy_comparison.png"
    fig.savefig(accuracy_path, dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10.5, 5.5))
    ax.bar(labels, correct, label="Correct", color="#4f8f5b")
    ax.bar(labels, incorrect, bottom=correct, label="Incorrect", color="#c85454")
    ax.set_ylabel("Images")
    ax.set_title("Updated 38-Image Test: Correct vs Incorrect")
    ax.set_xticklabels(labels, rotation=25, ha="right")
    ax.legend()
    fig.tight_layout()
    correctness_path = ASSET_DIR / "correct_vs_incorrect.png"
    fig.savefig(correctness_path, dpi=180)
    plt.close(fig)

    coral_acc = [r["coral_accuracy"] * 100 for r in per_class]
    shell_acc = [r["shell_accuracy"] * 100 for r in per_class]
    fig, ax = plt.subplots(figsize=(10.5, 5.5))
    ax.plot(labels, coral_acc, marker="o", linewidth=2.5, label="Corals", color="#9567a6")
    ax.plot(labels, shell_acc, marker="o", linewidth=2.5, label="Shells", color="#b58243")
    ax.set_ylabel("Class accuracy (%)")
    ax.set_title("Updated Test Accuracy by Class")
    ax.set_ylim(0, 105)
    ax.set_xticklabels(labels, rotation=25, ha="right")
    ax.legend()
    fig.tight_layout()
    per_class_path = ASSET_DIR / "per_class_accuracy.png"
    fig.savefig(per_class_path, dpi=180)
    plt.close(fig)

    return {
        "accuracy": accuracy_path,
        "correctness": correctness_path,
        "per_class": per_class_path,
    }


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2E3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 42)


def add_table(ws, name: str):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_rows(ws, headers: list[str], rows: list[dict]):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def create_workbook():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_tests, preprocessing, per_class, incorrect = collect_results()
    training = collect_training_history()
    chart_paths = make_charts(preprocessing, per_class)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Fossil Classifier Experiment Results"
    dash["A1"].font = Font(size=18, bold=True, color="1F4E5F")
    dash["A2"] = "Summary of exploratory tests, preprocessing shootout results, and updated 38-image wild test."
    dash["A4"] = "Current best model"
    dash["B4"] = "Original + enhanced"
    dash["A5"] = "Updated wild accuracy"
    dash["B5"] = 0.8947368421052632
    dash["B5"].number_format = "0.0%"
    dash["A6"] = "Updated correct / total"
    dash["B6"] = "34 / 38"
    dash["A7"] = "Important note"
    dash["B7"] = "The 38-image test is more reliable than the earlier 10-image test."
    for cell in ["A4", "A5", "A6", "A7"]:
        dash[cell].font = Font(bold=True)
    dash.column_dimensions["A"].width = 24
    dash.column_dimensions["B"].width = 80
    dash.add_image(XLImage(str(chart_paths["accuracy"])), "A10")
    dash.add_image(XLImage(str(chart_paths["correctness"])), "A40")

    ws = wb.create_sheet("All Tests")
    headers = ["test_id", "display_name", "family", "train_variants", "test_set", "total", "correct", "incorrect", "accuracy", "source_file"]
    write_rows(ws, headers, all_tests)
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        row[0].number_format = "0.0%"
    style_sheet(ws)
    add_table(ws, "AllTests")

    ws = wb.create_sheet("Model Comparison")
    headers = [
        "experiment",
        "display_name",
        "train_variants",
        "initial_correct",
        "initial_total",
        "initial_incorrect",
        "initial_accuracy",
        "updated_correct",
        "updated_total",
        "updated_incorrect",
        "updated_accuracy",
        "accuracy_change",
    ]
    write_rows(ws, headers, preprocessing)
    for col in ["G", "K", "L"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.0%"
    style_sheet(ws)
    add_table(ws, "ModelComparison")

    chart = BarChart()
    chart.title = "Initial vs Updated Wild Accuracy"
    chart.y_axis.title = "Accuracy"
    chart.x_axis.title = "Model"
    data = Reference(ws, min_col=7, max_col=11, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 20
    ws.add_chart(chart, "N2")

    ws = wb.create_sheet("Per Class Updated")
    headers = ["experiment", "coral_correct", "coral_total", "coral_accuracy", "shell_correct", "shell_total", "shell_accuracy", "coral_as_shell", "shell_as_coral"]
    write_rows(ws, headers, per_class)
    for col in ["D", "G"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.0%"
    style_sheet(ws)
    add_table(ws, "PerClassUpdated")
    ws.add_image(XLImage(str(chart_paths["per_class"])), "K2")

    ws = wb.create_sheet("Incorrect Updated")
    headers = ["experiment", "folder", "image", "true_class", "predicted_class", "confidence"]
    write_rows(ws, headers, incorrect)
    for cell in ws["F"][1:]:
        cell.number_format = "0.0%"
    style_sheet(ws)
    add_table(ws, "IncorrectUpdated")

    ws = wb.create_sheet("Training History")
    headers = ["model", "epoch", "train_accuracy", "val_accuracy", "train_loss", "val_loss"]
    write_rows(ws, headers, training)
    for col in ["C", "D"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.0%"
    for col in ["E", "F"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.000"
    style_sheet(ws)
    add_table(ws, "TrainingHistory")

    ws = wb.create_sheet("Notes")
    notes = [
        ("Term", "Meaning"),
        ("Initial wild test", "The earlier 10-image test set. Some exploratory tests evaluated original+enhanced variants, giving 20 evaluated samples."),
        ("Updated wild test", "The current 38-image test set: 25 coral images and 13 shell images."),
        ("Original + enhanced", "Training set contains both the original converted JPEG and an enhanced JPEG version of each training photo."),
        ("Grayscale / CLAHE", "Saved preprocessing variants used for experiment comparison; updated results suggest they do not currently improve generalization."),
        ("Validation accuracy", "Internal validation scores were often perfect, so the external wild test is more meaningful for choosing a model."),
    ]
    for row in notes:
        ws.append(row)
    style_sheet(ws)
    add_table(ws, "NotesTable")

    wb.save(OUTPUT_XLSX)


def verify_workbook():
    wb = load_workbook(OUTPUT_XLSX, data_only=False)
    expected = {
        "Dashboard",
        "All Tests",
        "Model Comparison",
        "Per Class Updated",
        "Incorrect Updated",
        "Training History",
        "Notes",
    }
    missing = expected - set(wb.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {missing}")
    if wb["Model Comparison"].max_row < 6:
        raise RuntimeError("Model Comparison sheet is unexpectedly small")
    if wb["Incorrect Updated"].max_row < 2:
        raise RuntimeError("Incorrect Updated sheet has no rows")
    print(f"Verified workbook: {OUTPUT_XLSX}")


if __name__ == "__main__":
    create_workbook()
    verify_workbook()
